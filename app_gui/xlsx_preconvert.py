"""Pre-convert staged XLSX files into AI-friendly normalized text assets."""

from dataclasses import dataclass, field
from datetime import date, datetime, time
import csv
import json
import os
import re
from typing import Dict, Iterable, List

from openpyxl import load_workbook


_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]+")
_SAMPLE_LIMIT = 12
_SUPPORTED_EXT = ".xlsx"


@dataclass
class XlsxPreconvertResult:
    """Summary payload for one pre-conversion run."""

    status: str = "skipped"
    message: str = ""
    normalized_root: str = ""
    normalized_assets: List[str] = field(default_factory=list)
    report_path: str = ""
    report: Dict[str, object] = field(default_factory=dict)


class XlsxPreconvertService:
    """Convert staged XLSX files into normalized CSV + schema summaries."""

    def convert_staged_files(self, source_paths: Iterable[str], *, normalized_root: str) -> XlsxPreconvertResult:
        normalized_root_abs = os.path.abspath(str(normalized_root or "").strip())
        if not normalized_root_abs:
            raise ValueError("normalized_root is required")
        os.makedirs(normalized_root_abs, exist_ok=True)

        staged_paths = _normalize_paths(source_paths)
        xlsx_paths = [path for path in staged_paths if str(path).lower().endswith(_SUPPORTED_EXT)]
        source_reports = []
        generated_assets = []

        for source_path in xlsx_paths:
            try:
                report = self._convert_one_xlsx(source_path, normalized_root_abs)
            except Exception as exc:  # pragma: no cover - defensive fallback
                report = {
                    "source_path": source_path,
                    "status": "failed",
                    "message": f"Failed to convert xlsx: {exc}",
                    "generated_files": [],
                    "warnings": [],
                    "errors": [{"message": str(exc)}],
                    "sheet_count": 0,
                }
            source_reports.append(report)
            generated_assets.extend(report.get("generated_files") or [])

        status = _merge_status(source_reports, xlsx_count=len(xlsx_paths))
        if status == "ok":
            message = "XLSX pre-conversion succeeded."
        elif status == "partial":
            message = "XLSX pre-conversion partially succeeded; fallback to raw files if needed."
        elif status == "failed":
            message = "XLSX pre-conversion failed; fallback to staged raw files."
        else:
            message = "No .xlsx files detected; staged files are kept as-is."

        report_payload = {
            "status": status,
            "message": message,
            "normalized_root": normalized_root_abs,
            "source_count": len(staged_paths),
            "xlsx_detected": len(xlsx_paths),
            "generated_assets": _dedupe_preserve_order(generated_assets),
            "sources": source_reports,
            "generated_at": _utc_timestamp(),
        }
        report_path = os.path.join(normalized_root_abs, "conversion_report.json")
        _write_json(report_path, report_payload)

        normalized_assets = _dedupe_preserve_order(generated_assets + [report_path])
        return XlsxPreconvertResult(
            status=status,
            message=message,
            normalized_root=normalized_root_abs,
            normalized_assets=normalized_assets,
            report_path=report_path,
            report=report_payload,
        )

    def _convert_one_xlsx(self, source_path: str, normalized_root: str) -> Dict[str, object]:
        source_abs = os.path.abspath(source_path)
        source_name = os.path.basename(source_abs)
        source_stem = _safe_name(os.path.splitext(source_name)[0], fallback="source")
        source_dir = _dedupe_dir(os.path.join(normalized_root, source_stem))
        sheets_dir = os.path.join(source_dir, "sheets")
        os.makedirs(sheets_dir, exist_ok=False)

        warnings = []
        errors = []
        generated_files = []
        sheet_summaries = []

        workbook = load_workbook(source_abs, read_only=True, data_only=True)
        try:
            sheet_names = list(workbook.sheetnames or [])
            if not sheet_names:
                raise ValueError("workbook has no worksheets")

            for sheet_index, sheet_name in enumerate(sheet_names, start=1):
                safe_sheet_name = _safe_name(sheet_name, fallback=f"sheet_{sheet_index}")
                csv_name = f"{sheet_index:02d}_{safe_sheet_name}.csv"
                csv_path = os.path.join(sheets_dir, csv_name)
                try:
                    sheet_summary = _write_sheet_csv_and_summary(
                        workbook[sheet_name],
                        csv_path,
                        sheet_name=sheet_name,
                        sheet_index=sheet_index,
                    )
                except Exception as exc:
                    errors.append(
                        {
                            "sheet_name": sheet_name,
                            "sheet_index": sheet_index,
                            "message": str(exc),
                        }
                    )
                    continue

                sheet_summary["csv_path"] = csv_path
                sheet_summaries.append(sheet_summary)
                generated_files.append(csv_path)
        finally:
            workbook.close()

        if errors and sheet_summaries:
            status = "partial"
        elif errors:
            status = "failed"
        else:
            status = "ok"

        schema_summary_payload = {
            "source_path": source_abs,
            "source_name": source_name,
            "status": status,
            "sheet_count": len(sheet_summaries),
            "sheets": sheet_summaries,
            "generated_at": _utc_timestamp(),
        }
        schema_summary_path = os.path.join(source_dir, "schema_summary.json")
        _write_json(schema_summary_path, schema_summary_payload)
        generated_files.append(schema_summary_path)

        report_payload = {
            "source_path": source_abs,
            "source_name": source_name,
            "status": status,
            "message": _status_message(status),
            "sheet_count": len(sheet_summaries),
            "schema_summary_path": schema_summary_path,
            "generated_files": list(generated_files),
            "warnings": warnings,
            "errors": errors,
            "generated_at": _utc_timestamp(),
        }
        report_path = os.path.join(source_dir, "conversion_report.json")
        _write_json(report_path, report_payload)
        generated_files.append(report_path)
        report_payload["report_path"] = report_path
        report_payload["generated_files"] = generated_files
        return report_payload


def _status_message(status: str) -> str:
    if status == "ok":
        return "Workbook converted to normalized CSV assets."
    if status == "partial":
        return "Workbook partially converted; review missing/error sheets."
    return "Workbook conversion failed; use staged raw file as fallback."


def _normalize_paths(paths: Iterable[str]) -> List[str]:
    normalized = []
    seen = set()
    for raw in paths or []:
        path = os.path.abspath(str(raw or "").strip())
        if not path:
            continue
        key = os.path.normcase(path)
        if key in seen:
            continue
        seen.add(key)
        normalized.append(path)
    return normalized


def _dedupe_preserve_order(paths: Iterable[str]) -> List[str]:
    deduped = []
    seen = set()
    for item in paths or []:
        path = os.path.abspath(str(item or "").strip())
        if not path:
            continue
        key = os.path.normcase(path)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(path)
    return deduped


def _merge_status(source_reports: List[Dict[str, object]], *, xlsx_count: int) -> str:
    if xlsx_count <= 0:
        return "skipped"
    statuses = [str(report.get("status") or "").strip().lower() for report in source_reports]
    success_count = sum(1 for status in statuses if status == "ok")
    partial_count = sum(1 for status in statuses if status == "partial")
    if success_count == xlsx_count:
        return "ok"
    if success_count > 0 or partial_count > 0:
        return "partial"
    return "failed"


def _safe_name(raw_name: str, *, fallback: str) -> str:
    sanitized = _SAFE_NAME_RE.sub("_", str(raw_name or "").strip()).strip("._")
    if not sanitized:
        return str(fallback)
    if len(sanitized) > 80:
        return sanitized[:80]
    return sanitized


def _dedupe_dir(base_path: str) -> str:
    candidate = os.path.abspath(base_path)
    if not os.path.exists(candidate):
        return candidate
    idx = 2
    while True:
        suffix = f"_{idx}"
        trial = f"{candidate}{suffix}"
        if not os.path.exists(trial):
            return trial
        idx += 1


def _write_sheet_csv_and_summary(worksheet, output_csv_path: str, *, sheet_name: str, sheet_index: int) -> Dict[str, object]:
    row_count = 0
    max_columns = 0
    non_empty_rows = 0
    header_candidate = None
    column_stats: List[Dict[str, object]] = []

    with open(output_csv_path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        for row_index, row_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            raw_row = _trim_trailing_none(list(row_values))
            text_row = [_to_text(value) for value in raw_row]
            writer.writerow(text_row)
            row_count += 1
            max_columns = max(max_columns, len(text_row))

            non_empty_values = [value for value in text_row if str(value).strip()]
            if non_empty_values:
                non_empty_rows += 1
                if header_candidate is None:
                    header_candidate = {
                        "row_index": row_index,
                        "values": list(text_row),
                    }

            _update_column_stats(column_stats, raw_row, text_row)

    return {
        "sheet_name": sheet_name,
        "sheet_index": sheet_index,
        "row_count": row_count,
        "non_empty_rows": non_empty_rows,
        "max_columns": max_columns,
        "header_candidate": header_candidate,
        "columns": _finalize_column_stats(column_stats),
    }


def _trim_trailing_none(values: List[object]) -> List[object]:
    trimmed = list(values)
    while trimmed and trimmed[-1] is None:
        trimmed.pop()
    return trimmed


def _to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        normalized = value.replace(microsecond=0)
        if (
            normalized.hour == 0
            and normalized.minute == 0
            and normalized.second == 0
            and normalized.tzinfo is None
        ):
            return normalized.date().isoformat()
        return normalized.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        normalized = value.replace(microsecond=0)
        return normalized.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return str(value)


def _infer_type(value: object) -> str:
    if value is None:
        return "empty"
    if isinstance(value, str) and not value.strip():
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    return "string"


def _update_column_stats(column_stats: List[Dict[str, object]], raw_row: List[object], text_row: List[str]) -> None:
    for index, raw_value in enumerate(raw_row, start=1):
        while len(column_stats) < index:
            column_stats.append(
                {
                    "index": len(column_stats) + 1,
                    "non_empty": 0,
                    "type_counts": {},
                    "sample_values": [],
                    "_sample_seen": set(),
                }
            )
        col = column_stats[index - 1]
        value_type = _infer_type(raw_value)
        col["type_counts"][value_type] = int(col["type_counts"].get(value_type, 0)) + 1
        if value_type != "empty":
            col["non_empty"] = int(col.get("non_empty", 0)) + 1
            sample_text = text_row[index - 1] if (index - 1) < len(text_row) else ""
            seen = col["_sample_seen"]
            if sample_text not in seen and len(col["sample_values"]) < _SAMPLE_LIMIT:
                col["sample_values"].append(sample_text)
                seen.add(sample_text)


def _finalize_column_stats(column_stats: List[Dict[str, object]]) -> List[Dict[str, object]]:
    finalized = []
    for col in column_stats:
        type_counts = dict(col.get("type_counts") or {})
        finalized.append(
            {
                "index": int(col.get("index") or 0),
                "non_empty": int(col.get("non_empty") or 0),
                "detected_type": _dominant_type(type_counts),
                "type_counts": type_counts,
                "sample_values": list(col.get("sample_values") or []),
            }
        )
    return finalized


def _dominant_type(type_counts: Dict[str, int]) -> str:
    non_empty_counts = {
        key: int(value)
        for key, value in (type_counts or {}).items()
        if key != "empty" and int(value) > 0
    }
    if not non_empty_counts:
        return "empty"
    rank = ["string", "datetime", "date", "integer", "number", "boolean", "time"]
    return sorted(non_empty_counts.items(), key=lambda item: (-item[1], rank.index(item[0]) if item[0] in rank else 99))[0][0]


def _write_json(path: str, payload: Dict[str, object]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def _utc_timestamp() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
